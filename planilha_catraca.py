import openpyxl

# 1. Carrega o arquivo Excel original (substitua pelo nome do seu arquivo)
nome_arquivo_original = "seu_arquivo.xlsx"
nome_arquivo_salvar = "seu_arquivo_tratado.xlsx"

wb = openpyxl.load_workbook(nome_arquivo_original)
ws = wb.active  # Seleciona a aba ativa (ou use wb["Nome da Aba"])

# 2. Percorre as linhas a partir da segunda (linha 1 são os cabeçalhos)
for row in range(2, ws.max_row + 1):
    
    # Tratando a coluna CodAluno (Coluna 1)
    celula_cod = ws.cell(row=row, column=1)
    val_cod = celula_cod.value
    if val_cod and str(val_cod).upper().startswith('M'):
        # Mantém apenas os caracteres após o primeiro (remove o 'M')
        celula_cod.value = str(val_cod)[1:]
        
    # Tratando a coluna NumCartao (Coluna 3)
    celula_cartao = ws.cell(row=row, column=3)
    val_cartao = celula_cartao.value
    if val_cartao and str(val_cartao).upper().startswith('M'):
        # Mantém apenas os caracteres após o primeiro (remove o 'M')
        celula_cartao.value = str(val_cartao)[1:]

# 3. Salva as alterações em um novo arquivo
wb.save(nome_arquivo_salvar)
print(f"Processo concluído! Arquivo salvo como: {nome_arquivo_salvar}")